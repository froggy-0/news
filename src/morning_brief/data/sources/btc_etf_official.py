from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import logging
import re
import zipfile
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from xml.etree import ElementTree as ET

# Re-export `date` so tests can use `official.date(...)` without a direct import.
__all_date__ = date  # noqa: F841 — kept for test compatibility

from perplexity import (
    APIConnectionError,
    APIStatusError,
    APITimeoutError,
    Perplexity,
    RateLimitError,
)

from morning_brief.data.etf_storage import EtfStorageBundle, build_storage_bundle_from_env
from morning_brief.data.sources.domain_utils import domain_matches, normalize_domain
from morning_brief.data.sources.http_client import HttpFetchError, get_bytes_with_retry
from morning_brief.data.sources.provider_runtime import (
    disabled_reason,
    execute_with_provider_retry,
    open_circuit,
    parse_retry_after_seconds,
    policy_for,
    record_skip,
)
from morning_brief.logging_utils import get_log_context, log_structured
from morning_brief.models import BitcoinEtfIssuerSnapshot
from morning_brief.observability import PipelineObserver

logger = logging.getLogger(__name__)

IBIT_URL = "https://www.ishares.com/us/products/333011/ishares-bitcoin-trust-etf"
BITB_URL = "https://bitbetf.com/"
GBTC_URL = "https://etfs.grayscale.com/gbtc"
BTC_MINI_URL = "https://etfs.grayscale.com/btc"
GBTC_XLSX_URL = (
    "https://reporting-prod-20231113144948145500000003.s3.us-east-1.amazonaws.com"
    "/product-performance/672e88c7-dac6-4fcd-9069-18eef01a2c73.xlsx"
)
BTC_MINI_XLSX_URL = (
    "https://reporting-prod-20231113144948145500000003.s3.us-east-1.amazonaws.com"
    "/product-performance/9ba286d6-3067-4153-b430-81d9d7a25696.xlsx"
)
IBIT_CREATION_BASKET_SHARES = 40_000
PERPLEXITY_PROVIDER = "perplexity"
OFFICIAL_BTC_ETF_PROVIDER = "official_btc_etf"
OFFICIAL_ETF_SCHEMA_VERSION = "v1"
BTC_ETF_REFERENCE_MODEL = "sonar"
BTC_ETF_ISSUER_DOMAIN_WHITELIST = (
    "ishares.com",
    "bitbetf.com",
    "etfs.grayscale.com",
)
BTC_ETF_ALLOWED_SOURCE_URLS = frozenset(
    {
        GBTC_XLSX_URL,
        BTC_MINI_XLSX_URL,
    }
)
BTC_ETF_REFERENCE_DOMAINS = BTC_ETF_ISSUER_DOMAIN_WHITELIST
BTC_ETF_REFERENCE_PROMPT = """
Return only JSON with this exact schema:
{
  "snapshots": [
    {
      "ticker": "IBIT",
      "issuer": "iShares",
      "source_url": "https://...",
      "as_of": "MM/DD/YYYY",
      "shares_outstanding": 0,
      "daily_volume": 0,
      "aum_usd": 0,
      "total_btc": 0,
      "bitcoin_per_share": 0
    }
  ]
}

Task:
- Find the latest available official spot Bitcoin ETF issuer data for IBIT, BITB, and GBTC.
- Use only official issuer domains.
- Include an item only if you can provide all numeric fields and a direct official source URL.
- Prefer the most recent daily snapshot available as of {today}.
- Do not include commentary, markdown, or citations outside the JSON.
""".strip()

BTC_ETF_REFERENCE_RESPONSE_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "snapshots": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "ticker": {"type": "string"},
                    "issuer": {"type": "string"},
                    "source_url": {"type": "string"},
                    "as_of": {"type": "string"},
                    "shares_outstanding": {"type": "number"},
                    "daily_volume": {"type": "number"},
                    "aum_usd": {"type": "number"},
                    "total_btc": {"type": "number"},
                    "bitcoin_per_share": {"type": "number"},
                },
                "required": [
                    "ticker",
                    "issuer",
                    "source_url",
                    "as_of",
                    "shares_outstanding",
                    "daily_volume",
                    "aum_usd",
                    "total_btc",
                ],
            },
        }
    },
    "required": ["snapshots"],
}

DATE_RE = r"(?:[A-Z][a-z]{2}\s+\d{1,2},\s+\d{4}|\d{2}/\d{2}/\d{4})"
VALUE_RE = r"\$?[\d,]+(?:\.\d+)?(?:[MB])?"
NEXT_DATA_RE = re.compile(
    r'<script[^>]+id="__NEXT_DATA__"[^>]*>\s*(?P<payload>\{.*?\})\s*</script>',
    flags=re.DOTALL | re.IGNORECASE,
)
JSON_CODE_BLOCK_RE = re.compile(
    r"```(?:json)?\s*(?P<payload>\{.*?\})\s*```",
    flags=re.DOTALL | re.IGNORECASE,
)
RESPONSE_PREVIEW_LEN = 200
OFFICIAL_BTC_ETF_STATE_FILE = "state.json"
DOWNLOAD_LINK_RE = re.compile(
    r"""(?P<attr>href|src)=["'](?P<url>[^"']+\.(?:csv|json|xlsx))["']""",
    flags=re.IGNORECASE,
)
HTML_ONLY_SOURCE_TYPE = "official_html"
STRUCTURED_SOURCE_PRIORITY = {
    "IBIT": ("official_csv", "official_json", "official_html"),
    "BITB": ("official_json", "official_csv", "official_html"),
    "GBTC": ("official_csv",),
    "BTC": ("official_csv",),
}


@dataclass(frozen=True)
class CollectedOfficialSnapshot:
    snapshot: BitcoinEtfIssuerSnapshot
    raw_payload: bytes
    source_format: str
    parse_method: str
    source_file_url: str | None = None
    raw_label_map: dict[str, str] | None = None
    raw_text_map: dict[str, str] | None = None
    http_status: int = 200
    reference_only: bool = False


def _parse_as_of_date(raw: str) -> date:
    """Parse various as_of date string formats into a date object."""
    raw = raw.strip()
    for fmt in ("%m/%d/%Y", "%b %d, %Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized as_of date format: {raw!r}")


def _current_run_id(observer: PipelineObserver | None) -> str:
    if observer is not None:
        return observer.run_id
    context_run_id = get_log_context().get("run_id")
    if isinstance(context_run_id, str) and context_run_id:
        return context_run_id
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _sha256_hex(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _compute_quality_status(source_type: str) -> str:
    """Req 8: ok for json/csv, degraded for html."""
    if source_type in ("official_json", "official_csv"):
        return "ok"
    if source_type == "aggregator":
        return "degraded"
    return "degraded"


def _normalize_source_format(url: str, source_type: str) -> str:
    lowered = url.lower()
    if lowered.endswith(".json"):
        return "json"
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    if source_type == "official_json":
        return "json"
    if source_type == "official_csv":
        return "csv"
    return "html"


def _discover_download_links(page_text: str, *, base_url: str) -> list[str]:
    links: list[str] = []
    for match in DOWNLOAD_LINK_RE.finditer(page_text):
        candidate = match.group("url").strip()
        if not candidate:
            continue
        absolute = urljoin(base_url, candidate)
        if absolute not in links:
            links.append(absolute)
    return links


def _allowed_source_url(url: str) -> bool:
    if url.strip() in BTC_ETF_ALLOWED_SOURCE_URLS:
        return True
    domain = normalize_domain(url)
    return any(domain_matches(domain, candidate) for candidate in BTC_ETF_ISSUER_DOMAIN_WHITELIST)


def _validate_issuer_url(url: str) -> None:
    if not _allowed_source_url(url):
        raise ValueError("필수 ETF 참조 필드가 없거나 공식 도메인이 아니에요.")


def _xlsx_cell_value(
    cell: ET.Element,
    *,
    shared_strings: list[str],
) -> str:
    value_node = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
    inline_node = cell.find(
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is/"
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"
    )
    if inline_node is not None and inline_node.text is not None:
        return inline_node.text.strip()
    if value_node is None or value_node.text is None:
        return ""
    raw = value_node.text.strip()
    if cell.get("t") == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return ""
    return raw


def _xlsx_column_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha()).upper()
    value = 0
    for ch in letters:
        value = value * 26 + (ord(ch) - 64)
    return max(value - 1, 0)


def _read_xlsx_rows(payload: bytes) -> list[list[str]]:
    workbook_rows = _read_xlsx_workbook_rows(payload)
    if not workbook_rows:
        return []
    return next(iter(workbook_rows.values()))


def _read_xlsx_workbook_rows(payload: bytes) -> dict[str, list[list[str]]]:
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        openpyxl_rows: dict[str, list[list[str]]] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                sheet_rows.append(["" if value is None else str(value).strip() for value in row])
            openpyxl_rows[sheet_name] = sheet_rows
        return openpyxl_rows
    except Exception:
        pass

    archive = zipfile.ZipFile(io.BytesIO(payload))
    shared_strings: list[str] = []
    if "xl/sharedStrings.xml" in archive.namelist():
        shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
        shared_strings = [
            "".join(node.itertext()).strip()
            for node in shared_root.findall(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"
            )
        ]

    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
        for rel in rels_root
        if rel.attrib.get("Id")
    }
    workbook_rows: dict[str, list[list[str]]] = {}
    for sheet in workbook_root.findall(
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/"
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
    ):
        sheet_name = sheet.attrib.get("name", "").strip()
        rel_id = sheet.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        )
        if not sheet_name or not rel_id:
            continue
        target = rel_targets.get(rel_id, "").strip()
        if not target:
            continue
        sheet_path = target if target.startswith("xl/") else f"xl/{target}"
        sheet_root = ET.fromstring(archive.read(sheet_path))
        rows: list[list[str]] = []
        for row in sheet_root.findall(
            ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row"
        ):
            values: list[str] = []
            for cell in row.findall("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"):
                index = _xlsx_column_index(cell.attrib.get("r", "A1"))
                while len(values) <= index:
                    values.append("")
                values[index] = _xlsx_cell_value(cell, shared_strings=shared_strings)
            rows.append(values)
        workbook_rows[sheet_name] = rows
    return workbook_rows


def _header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _select_xlsx_data_row(rows: list[list[str]]) -> tuple[list[str], dict[str, str]]:
    header_index = -1
    headers: list[str] = []
    for idx, row in enumerate(rows):
        normalized_headers = {_header_key(cell) for cell in row if cell.strip()}
        if "asofdate" in normalized_headers or "date" in normalized_headers:
            header_index = idx
            headers = row
            break
    if header_index == -1 or not headers:
        raise HttpFetchError("구조화 다운로드에서 헤더 행을 찾지 못했어요.")

    normalized_headers = [_header_key(header) for header in headers]
    date_keys = [candidate for candidate in ("asofdate", "date") if candidate in normalized_headers]
    if not date_keys:
        raise HttpFetchError("구조화 다운로드에서 기준일 컬럼을 찾지 못했어요.")
    best_date: date | None = None
    best_row: dict[str, str] | None = None
    for row in rows[header_index + 1 :]:
        record = {
            normalized_headers[idx]: row[idx].strip()
            for idx in range(min(len(normalized_headers), len(row)))
            if normalized_headers[idx]
        }
        as_of_raw = next((record.get(key, "") for key in date_keys if record.get(key, "")), "")
        if not as_of_raw:
            continue
        try:
            current_date = _parse_as_of_date(as_of_raw)
        except ValueError:
            try:
                current_date = datetime.fromisoformat(as_of_raw).date()
            except ValueError:
                continue
        if best_date is None or current_date > best_date:
            best_date = current_date
            best_row = record
    if best_row is None:
        raise HttpFetchError("구조화 다운로드에서 기준일 레코드를 찾지 못했어요.")
    return headers, best_row


def _first_record_value(record: dict[str, str], candidates: list[str]) -> str:
    for candidate in candidates:
        value = record.get(_header_key(candidate), "").strip()
        if value:
            return value
    raise HttpFetchError("구조화 다운로드에서 필수 필드를 찾지 못했어요.")


def _validate_snapshot_anomalies(
    snapshot: "BitcoinEtfIssuerSnapshot",
) -> "BitcoinEtfIssuerSnapshot":
    """Req 9: apply anomaly validation rules, escalate quality_status when needed."""
    from dataclasses import replace as _replace

    updated = snapshot
    # Req 9.1: shares_outstanding <= 0
    if updated.shares_outstanding <= 0:
        log_structured(
            logger,
            event="etf.anomaly_invalid_field",
            message="유효하지 않은 shares_outstanding 값을 감지했어요.",
            level=logging.WARNING,
            ticker=updated.ticker,
            field="shares_outstanding",
            value=updated.shares_outstanding,
        )
        updated = _replace(updated, quality_status="critical")
    # Req 9.2: total_btc < 0
    if updated.total_btc < 0:
        log_structured(
            logger,
            event="etf.anomaly_invalid_field",
            message="유효하지 않은 total_btc 값을 감지했어요.",
            level=logging.WARNING,
            ticker=updated.ticker,
            field="total_btc",
            value=updated.total_btc,
        )
        updated = _replace(updated, quality_status="critical")
    return updated


def _render_reference_prompt(today: date) -> str:
    return BTC_ETF_REFERENCE_PROMPT.replace("{today}", today.isoformat())


def _normalize_page_text(text: str) -> str:
    cleaned = re.sub(r"(?is)<script.*?>.*?</script>", " ", text)
    cleaned = re.sub(r"(?is)<style.*?>.*?</style>", " ", cleaned)
    cleaned = re.sub(r"(?s)<[^>]+>", " ", cleaned)
    cleaned = html.unescape(cleaned)
    cleaned = cleaned.replace("\xa0", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip()


def _parse_compact_number(raw: str) -> float:
    normalized = raw.strip().replace("$", "").replace(",", "")
    multiplier = 1.0
    if normalized.endswith("B"):
        multiplier = 1_000_000_000.0
        normalized = normalized[:-1]
    elif normalized.endswith("M"):
        multiplier = 1_000_000.0
        normalized = normalized[:-1]
    return float(normalized) * multiplier


def _coerce_float(value: object) -> float | None:
    if not isinstance(value, (int, float, str, bytes)):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _extract_value(text: str, label: str) -> float:
    match = re.search(rf"{re.escape(label)}\*?\s+(?P<value>{VALUE_RE})", text, flags=re.IGNORECASE)
    if not match:
        raise HttpFetchError(f"공식 ETF 페이지에서 '{label}' 값을 찾지 못했어요.")
    return _parse_compact_number(match.group("value"))


def _extract_first_matching_value(text: str, labels: list[str]) -> float:
    last_error: Exception | None = None
    for label in labels:
        try:
            return _extract_value(text, label)
        except Exception as exc:
            last_error = exc
    raise HttpFetchError("공식 ETF 페이지에서 값을 찾지 못했어요.") from last_error


def _extract_optional_matching_value(text: str, labels: list[str]) -> float | None:
    try:
        return _extract_first_matching_value(text, labels)
    except Exception:
        return None


def _extract_dated_value(text: str, label: str) -> tuple[str, float]:
    match = re.search(
        rf"{re.escape(label)}\s+as of\s+(?P<date>{DATE_RE})\s+(?P<value>{VALUE_RE})",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        raise HttpFetchError(f"공식 ETF 페이지에서 날짜가 포함된 '{label}' 값을 찾지 못했어요.")
    return match.group("date"), _parse_compact_number(match.group("value"))


def _extract_page_date(text: str) -> str:
    match = re.search(rf"(?:Data|data)\s+as\s+of\s+(?P<date>{DATE_RE})", text)
    if not match:
        raise HttpFetchError("공식 ETF 페이지에서 기준일을 찾지 못했어요.")
    return match.group("date")


def _extract_next_data_payload(text: str) -> dict:
    match = NEXT_DATA_RE.search(text)
    if not match:
        raise HttpFetchError("Bitwise 페이지에서 __NEXT_DATA__를 찾지 못했어요.")

    try:
        payload = json.loads(match.group("payload"))
    except json.JSONDecodeError as exc:
        raise HttpFetchError("Bitwise 페이지의 __NEXT_DATA__ JSON을 읽지 못했어요.") from exc

    if not isinstance(payload, dict):
        raise HttpFetchError("Bitwise 페이지의 __NEXT_DATA__ 구조가 예상과 달라요.")
    return payload


def _find_first_key(payload: object, key: str) -> object | None:
    if isinstance(payload, dict):
        if key in payload:
            return payload[key]
        for value in payload.values():
            found = _find_first_key(value, key)
            if found is not None:
                return found
    elif isinstance(payload, list):
        for item in payload:
            found = _find_first_key(item, key)
            if found is not None:
                return found
    return None


def _extract_bitb_structured_optional_values(text: str) -> dict[str, float | None]:
    payload = _extract_next_data_payload(text)
    nav_and_market_price = _find_first_key(payload, "navAndMarketPrice")
    premium_discount = _find_first_key(payload, "premiumDiscount")
    optional_values: dict[str, float | None] = {
        "nav_per_share": None,
        "market_price": None,
        "premium_discount_pct": None,
    }

    if isinstance(nav_and_market_price, dict):
        optional_values["nav_per_share"] = _coerce_float(nav_and_market_price.get("nav"))
        optional_values["market_price"] = _coerce_float(nav_and_market_price.get("marketPrice"))

    if isinstance(premium_discount, dict):
        as_of_values = premium_discount.get("asOfValues")
        if isinstance(as_of_values, list) and as_of_values:
            latest = as_of_values[-1]
            if isinstance(latest, dict):
                optional_values["premium_discount_pct"] = _coerce_float(latest.get("value"))

    return optional_values


def _extract_bitb_structured_values(text: str) -> tuple[str, float, float, int, int]:
    payload = _extract_next_data_payload(text)
    as_of_raw = _find_first_key(payload, "timestamp")
    total_btc_raw = _find_first_key(payload, "totalReserve")
    aum_usd_raw = _find_first_key(payload, "netAssets")
    shares_outstanding_raw = _find_first_key(payload, "sharesOutstanding")
    daily_volume_raw = _find_first_key(payload, "volume")

    if not all(
        value is not None
        for value in [total_btc_raw, aum_usd_raw, shares_outstanding_raw, daily_volume_raw]
    ):
        raise HttpFetchError("Bitwise 페이지의 구조화 데이터가 충분하지 않아요.")

    as_of = _extract_page_date(_normalize_page_text(text))
    if isinstance(as_of_raw, str) and as_of_raw.strip():
        try:
            as_of = datetime.fromisoformat(as_of_raw.replace("Z", "+00:00")).strftime("%m/%d/%Y")
        except ValueError:
            as_of = as_of_raw[:10].replace("-", "/")

    return (
        as_of,
        float(_parse_numeric(total_btc_raw)),
        float(_parse_numeric(aum_usd_raw)),
        int(_parse_numeric(shares_outstanding_raw, integer=True)),
        int(_parse_numeric(daily_volume_raw, integer=True)),
    )


def _structured_download_links(page_text: str, *, base_url: str) -> list[tuple[str, str]]:
    links = _discover_download_links(page_text, base_url=base_url)
    structured_links: list[tuple[str, str]] = []
    for link in links:
        _validate_issuer_url(link)
        if link.lower().endswith(".json"):
            structured_links.append(("official_json", link))
        elif link.lower().endswith(".csv") or link.lower().endswith(".xlsx"):
            structured_links.append(("official_csv", link))
    return structured_links


def _normalize_record_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _parse_json_records(payload: bytes) -> list[dict[str, Any]]:
    decoded = json.loads(payload.decode("utf-8"))
    if isinstance(decoded, list):
        return [item for item in decoded if isinstance(item, dict)]
    if isinstance(decoded, dict):
        if all(not isinstance(value, (dict, list)) for value in decoded.values()):
            return [decoded]
        for key in ("data", "items", "rows", "results", "snapshots"):
            value = decoded.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
        for value in decoded.values():
            if isinstance(value, list):
                records = [item for item in value if isinstance(item, dict)]
                if records:
                    return records
    raise HttpFetchError("구조화 JSON에서 레코드 배열을 찾지 못했어요.")


def _parse_csv_records(payload: bytes) -> list[dict[str, Any]]:
    text = payload.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader if isinstance(row, dict)]


def _record_value(record: dict[str, Any], candidates: list[str]) -> Any | None:
    normalized = {_normalize_record_key(key): value for key, value in record.items()}
    for candidate in candidates:
        value = normalized.get(_normalize_record_key(candidate))
        if value not in (None, ""):
            return value
    return None


def _record_date(record: dict[str, Any]) -> date | None:
    raw_value = _record_value(
        record,
        [
            "as_of_date",
            "as_of",
            "date",
            "nav_date",
            "timestamp",
            "effective_date",
        ],
    )
    if raw_value in (None, ""):
        return None
    raw = str(raw_value).strip()
    if "T" in raw:
        raw = raw.split("T", 1)[0]
    try:
        return _parse_as_of_date(raw)
    except ValueError:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            return None


def _select_latest_structured_record(records: list[dict[str, Any]]) -> dict[str, Any]:
    best_record: dict[str, Any] | None = None
    best_date: date | None = None
    for record in records:
        current_date = _record_date(record)
        if current_date is None:
            continue
        if best_date is None or current_date > best_date:
            best_date = current_date
            best_record = record
    if best_record is None:
        raise HttpFetchError("구조화 다운로드에서 최신 기준일 레코드를 찾지 못했어요.")
    return best_record


def _structured_field_candidates(ticker: str) -> dict[str, list[str]]:
    base = {
        "as_of": ["as_of_date", "as_of", "date", "nav_date", "timestamp"],
        "shares_outstanding": ["shares_outstanding", "shares", "shares_out"],
        "daily_volume": ["daily_volume", "daily_volume_shares", "volume", "daily_shares_volume"],
        "aum_usd": [
            "aum_usd",
            "net_assets",
            "net_assets_aum",
            "assets_under_management",
            "assets_under_management_nongaap",
            "non_gaap_aum",
            "aum",
        ],
        "total_btc": [
            "total_btc",
            "bitcoin_holdings",
            "bitcoin_in_trust",
            "total_bitcoin_in_fund",
            "total_bitcoin_in_trust",
            "total_reserve",
        ],
        "bitcoin_per_share": ["bitcoin_per_share", "btc_per_share", "bitcoinshare"],
    }
    if ticker == "IBIT":
        base["basket_bitcoin_amount"] = ["basket_bitcoin_amount", "basket_bitcoin"]
    return base


def _parse_structured_snapshot(
    *,
    ticker: str,
    issuer: str,
    page_url: str,
    source_url: str,
    source_type: str,
    payload: bytes,
) -> BitcoinEtfIssuerSnapshot:
    records = (
        _parse_json_records(payload)
        if source_type == "official_json"
        else _parse_csv_records(payload)
    )
    record = _select_latest_structured_record(records)
    field_candidates = _structured_field_candidates(ticker)
    as_of_raw = _record_value(record, field_candidates["as_of"])
    shares_raw = _record_value(record, field_candidates["shares_outstanding"])
    aum_raw = _record_value(record, field_candidates["aum_usd"])
    volume_raw = _record_value(record, field_candidates["daily_volume"])
    total_btc_raw = _record_value(record, field_candidates["total_btc"])
    bitcoin_per_share_raw = _record_value(record, field_candidates["bitcoin_per_share"])

    if as_of_raw in (None, "") or shares_raw in (None, "") or aum_raw in (None, ""):
        raise HttpFetchError("구조화 다운로드에 필수 필드가 부족해요.")

    shares_outstanding = int(_parse_numeric(shares_raw, integer=True))
    daily_volume = (
        int(_parse_numeric(volume_raw, integer=True)) if volume_raw not in (None, "") else 0
    )
    aum_usd = float(_parse_numeric(aum_raw))
    total_btc: float | None = None
    bitcoin_per_share: float | None = None
    extra_fields: dict[str, Any] = {}

    if total_btc_raw not in (None, ""):
        total_btc = float(_parse_numeric(total_btc_raw))
    if bitcoin_per_share_raw not in (None, ""):
        bitcoin_per_share = float(_parse_numeric(bitcoin_per_share_raw))

    if ticker == "IBIT" and (total_btc is None or bitcoin_per_share is None):
        basket_raw = _record_value(record, field_candidates.get("basket_bitcoin_amount", []))
        if basket_raw not in (None, ""):
            basket_bitcoin_amount = float(_parse_numeric(basket_raw))
            bitcoin_per_share = basket_bitcoin_amount / IBIT_CREATION_BASKET_SHARES
            total_btc = shares_outstanding * bitcoin_per_share
            extra_fields["basket_bitcoin_amount"] = round(basket_bitcoin_amount, 8)

    if total_btc is None and bitcoin_per_share is not None:
        total_btc = shares_outstanding * bitcoin_per_share
        extra_fields["estimated_total_btc"] = round(total_btc, 8)
    if bitcoin_per_share is None and total_btc is not None:
        bitcoin_per_share = total_btc / shares_outstanding

    if total_btc is None or bitcoin_per_share is None:
        raise HttpFetchError("구조화 다운로드에서 BTC 보유량을 계산하지 못했어요.")

    snapshot = _snapshot_from_values(
        ticker=ticker,
        issuer=issuer,
        source_url=page_url,
        as_of_str=str(as_of_raw),
        shares_outstanding=shares_outstanding,
        daily_volume=daily_volume,
        aum_usd=aum_usd,
        total_btc=total_btc,
        bitcoin_per_share=bitcoin_per_share,
        source_type=source_type,
        extra_fields=extra_fields,
    )
    if "estimated_total_btc" in extra_fields:
        from dataclasses import replace as _replace

        snapshot = _replace(snapshot, quality_status="degraded")
    return snapshot


def _collect_structured_snapshot(
    *,
    ticker: str,
    issuer: str,
    page_url: str,
    source_type: str,
    source_url: str,
) -> CollectedOfficialSnapshot:
    payload = get_bytes_with_retry(
        source_url,
        provider=OFFICIAL_BTC_ETF_PROVIDER,
        headers={"Accept": "application/json,text/csv,application/octet-stream"},
    )
    if source_url.lower().endswith(".xlsx"):
        snapshot = _parse_grayscale_xlsx_snapshot(
            ticker=ticker,
            issuer=issuer,
            page_url=page_url,
            xlsx_url=source_url,
            payload=payload,
        )
        parse_method = f"{ticker.lower()}_xlsx"
    else:
        snapshot = _parse_structured_snapshot(
            ticker=ticker,
            issuer=issuer,
            page_url=page_url,
            source_url=source_url,
            source_type=source_type,
            payload=payload,
        )
        parse_method = f"{ticker.lower()}_{_normalize_source_format(source_url, source_type)}"
    return CollectedOfficialSnapshot(
        snapshot=snapshot,
        raw_payload=payload,
        source_format=_normalize_source_format(source_url, source_type),
        parse_method=parse_method,
        source_file_url=source_url,
        http_status=200,
    )


def _snapshot_from_values(
    *,
    ticker: str,
    issuer: str,
    source_url: str,
    as_of_str: str,
    shares_outstanding: int,
    daily_volume: int,
    aum_usd: float,
    total_btc: float,
    bitcoin_per_share: float,
    source_type: str,
    extra_fields: dict[str, Any] | None = None,
) -> BitcoinEtfIssuerSnapshot:
    snapshot = BitcoinEtfIssuerSnapshot(
        ticker=ticker,
        issuer=issuer,
        source_url=source_url,
        as_of_date=_parse_as_of_date(as_of_str),
        shares_outstanding=int(round(shares_outstanding)),
        daily_volume=int(round(daily_volume)),
        aum_usd=round(aum_usd, 2),
        total_btc=round(total_btc, 8),
        bitcoin_per_share=round(bitcoin_per_share, 10),
        source_type=source_type,
        quality_status=_compute_quality_status(source_type),
        collected_at=datetime.now(timezone.utc),
        extra_fields=extra_fields or {},
    )
    return _validate_snapshot_anomalies(snapshot)


def _parse_grayscale_xlsx_snapshot(
    *,
    ticker: str,
    issuer: str,
    page_url: str,
    xlsx_url: str,
    payload: bytes,
) -> BitcoinEtfIssuerSnapshot:
    _validate_issuer_url(xlsx_url)
    workbook_rows = _read_xlsx_workbook_rows(payload)
    daily_rows = workbook_rows.get("Daily Performance", _read_xlsx_rows(payload))
    _, record = _select_xlsx_data_row(daily_rows)
    as_of_str = _first_record_value(record, ["As of Date", "Date"])
    shares_outstanding = int(
        _parse_numeric(_first_record_value(record, ["Shares Outstanding"]), integer=True)
    )
    aum_usd = float(
        _parse_numeric(
            _first_record_value(
                record,
                ["Non-GAAP AUM", "AUM (Non-GAAP)", "Assets Under Management", "AUM"],
            )
        )
    )
    daily_volume_raw = _record_value(record, ["Daily Volume (Shares)"])
    daily_volume = (
        int(_parse_numeric(daily_volume_raw, integer=True))
        if daily_volume_raw not in (None, "")
        else 0
    )
    extra_fields = {}
    for field_name, candidates in {
        "nav_per_share": ["NAV per Share"],
        "market_price": ["Market Price", "Market Price Per Share"],
        "premium_discount_pct": ["Premium/Discount", "Premium Discount"],
        "gaap_aum": ["GAAP AUM"],
        "gaap_nav_per_share": ["GAAP NAV per Share"],
        "bid_ask_spread_30d": ["Bid Ask Spread 30D", "Bid/Ask Spread 30D"],
        "sponsor_fee": ["Sponsor Fee"],
    }.items():
        try:
            extra_fields[field_name] = float(
                _parse_numeric(_first_record_value(record, candidates))
            )
        except Exception:
            continue

    holdings_rows = workbook_rows.get("Holdings")
    if holdings_rows:
        _, holdings_record = _select_xlsx_data_row(holdings_rows)
        holdings_as_of_str = _first_record_value(holdings_record, ["As of Date", "Date"])
        if holdings_as_of_str:
            as_of_str = holdings_as_of_str
        bitcoin_per_share = float(
            _parse_numeric(
                _first_record_value(holdings_record, ["Asset/Share", "Bitcoin per Share"])
            )
        )
        total_btc = shares_outstanding * bitcoin_per_share
    else:
        total_btc = float(
            _parse_numeric(
                _first_record_value(
                    record, ["Total Bitcoin in Fund", "Bitcoin Holdings", "Total Bitcoin in Trust"]
                )
            )
        )
        bitcoin_per_share = float(
            _parse_numeric(_first_record_value(record, ["Bitcoin per Share"]))
        )

    snapshot = _snapshot_from_values(
        ticker=ticker,
        issuer=issuer,
        source_url=page_url,
        as_of_str=as_of_str,
        shares_outstanding=shares_outstanding,
        daily_volume=daily_volume,
        aum_usd=aum_usd,
        total_btc=total_btc,
        bitcoin_per_share=bitcoin_per_share,
        source_type="official_csv",
        extra_fields=extra_fields,
    )
    if daily_volume == 0 and holdings_rows:
        from dataclasses import replace as _replace

        snapshot = _replace(snapshot, quality_status="degraded")
    return snapshot


def parse_ibit_snapshot(text: str) -> BitcoinEtfIssuerSnapshot:
    normalized = _normalize_page_text(text)
    as_of_str, aum_usd = _extract_dated_value(normalized, "Net Assets of Fund")
    _, shares_outstanding = _extract_dated_value(normalized, "Shares Outstanding")
    _, daily_volume = _extract_dated_value(normalized, "Daily Volume")
    _, basket_bitcoin_amount = _extract_dated_value(normalized, "Basket Bitcoin Amount")
    bitcoin_per_share = basket_bitcoin_amount / IBIT_CREATION_BASKET_SHARES
    total_btc = bitcoin_per_share * shares_outstanding
    return _snapshot_from_values(
        ticker="IBIT",
        issuer="iShares",
        source_url=IBIT_URL,
        as_of_str=as_of_str,
        shares_outstanding=int(round(shares_outstanding)),
        daily_volume=int(round(daily_volume)),
        aum_usd=aum_usd,
        total_btc=total_btc,
        bitcoin_per_share=bitcoin_per_share,
        source_type=HTML_ONLY_SOURCE_TYPE,
        extra_fields={
            "basket_bitcoin_amount": round(basket_bitcoin_amount, 8),
            "closing_price": _extract_optional_matching_value(
                normalized, ["Closing Price", "Market Price", "Closing Market Price"]
            ),
            "premium_discount_pct": _extract_optional_matching_value(
                normalized, ["Premium/Discount", "Premium Discount"]
            ),
            "nav_per_share": _extract_optional_matching_value(normalized, ["NAV", "NAV per Share"]),
            "sponsor_fee": _extract_optional_matching_value(normalized, ["Sponsor Fee"]),
        },
    )


def parse_bitb_snapshot(text: str) -> BitcoinEtfIssuerSnapshot:
    normalized = _normalize_page_text(text)
    source_type: str
    optional_values = {
        "nav_per_share": None,
        "market_price": None,
        "premium_discount_pct": None,
    }
    try:
        as_of_str, total_btc, aum_usd, shares_outstanding, daily_volume = (
            _extract_bitb_structured_values(text)
        )
        bitcoin_per_share = total_btc / shares_outstanding
        source_type = "official_json"
        optional_values = _extract_bitb_structured_optional_values(text)
    except HttpFetchError:
        as_of_str = _extract_page_date(normalized)
        aum_usd = _extract_first_matching_value(normalized, ["Net Assets (AUM)", "Net Assets"])
        shares_outstanding = int(round(_extract_value(normalized, "Shares Outstanding")))
        daily_volume = int(
            round(
                _extract_first_matching_value(normalized, ["Daily Volume (Shares)", "Daily Volume"])
            )
        )
        total_btc = _extract_value(normalized, "Bitcoin in Trust")
        bitcoin_per_share = _extract_value(normalized, "Bitcoin per Share")
        source_type = "official_html"
    return _snapshot_from_values(
        ticker="BITB",
        issuer="Bitwise",
        source_url=BITB_URL,
        as_of_str=as_of_str,
        shares_outstanding=int(round(shares_outstanding)),
        daily_volume=int(round(daily_volume)),
        aum_usd=aum_usd,
        total_btc=total_btc,
        bitcoin_per_share=bitcoin_per_share,
        source_type=source_type,
        extra_fields={
            "nav_per_share": optional_values["nav_per_share"]
            if source_type == "official_json"
            else _extract_optional_matching_value(normalized, ["NAV"]),
            "market_price": optional_values["market_price"]
            if source_type == "official_json"
            else _extract_optional_matching_value(normalized, ["Market Price"]),
            "premium_discount_pct": optional_values["premium_discount_pct"]
            if source_type == "official_json"
            else _extract_optional_matching_value(
                normalized, ["Premium/Discount", "Premium Discount"]
            ),
            "bitcoin_reserve_btc": total_btc,
            "trust_net_assets_btc": total_btc,
            "bitcoin_in_trust": total_btc,
            "net_assets_aum": aum_usd,
        },
    )


def _parse_grayscale_snapshot(
    ticker: str, issuer: str, url: str, text: str
) -> BitcoinEtfIssuerSnapshot:
    """Req 6.4: Grayscale 공식 페이지 구조가 동일하므로 공용 파서를 재사용한다."""
    normalized = _normalize_page_text(text)
    as_of_str = _extract_page_date(normalized)
    aum_usd = _extract_first_matching_value(
        normalized,
        ["ASSETS UNDER MANAGEMENT (NON-GAAP)", "ASSETS UNDER MANAGEMENT"],
    )
    shares_outstanding = _extract_value(normalized, "SHARES OUTSTANDING")
    daily_volume = _extract_value(normalized, "DAILY VOLUME (SHARES)")
    total_btc = _extract_first_matching_value(
        normalized,
        ["TOTAL BITCOIN IN FUND", "TOTAL BITCOIN IN TRUST"],
    )
    bitcoin_per_share = _extract_value(normalized, "BITCOIN PER SHARE")
    return _snapshot_from_values(
        ticker=ticker,
        issuer=issuer,
        source_url=url,
        as_of_str=as_of_str,
        shares_outstanding=int(round(shares_outstanding)),
        daily_volume=int(round(daily_volume)),
        aum_usd=aum_usd,
        total_btc=total_btc,
        bitcoin_per_share=bitcoin_per_share,
        source_type=HTML_ONLY_SOURCE_TYPE,
        extra_fields={
            "aum_non_gaap": aum_usd,
            "nav_per_share": _extract_optional_matching_value(
                normalized, ["NAV PER SHARE", "GAAP NAV PER SHARE"]
            ),
            "market_price": _extract_optional_matching_value(normalized, ["MARKET PRICE"]),
            "premium_discount_pct": _extract_optional_matching_value(
                normalized, ["PREMIUM/DISCOUNT", "PREMIUM DISCOUNT"]
            ),
            "bid_ask_spread_30d": _extract_optional_matching_value(
                normalized, ["BID ASK SPREAD 30D", "BID/ASK SPREAD 30D"]
            ),
            "total_bitcoin_in_trust": total_btc,
        },
    )


def parse_gbtc_snapshot(text: str) -> BitcoinEtfIssuerSnapshot:
    return _parse_grayscale_snapshot("GBTC", "Grayscale", GBTC_URL, text)


def parse_btc_mini_snapshot(text: str) -> BitcoinEtfIssuerSnapshot:
    return _parse_grayscale_snapshot("BTC", "Grayscale Bitcoin Mini Trust", BTC_MINI_URL, text)


def _persist_collected_snapshot(
    bundle: EtfStorageBundle | None,
    *,
    run_id: str,
    collected: CollectedOfficialSnapshot,
) -> None:
    if bundle is None:
        return
    bundle.persist_snapshot(
        run_id=run_id,
        snapshot=collected.snapshot,
        source_format=collected.source_format,
        parse_method=collected.parse_method,
        payload=collected.raw_payload,
        source_checksum=_sha256_hex(collected.raw_payload),
        http_status=collected.http_status,
        source_file_url=collected.source_file_url,
        raw_label_map=collected.raw_label_map,
        raw_text_map=collected.raw_text_map,
        reference_only=collected.reference_only,
    )


def _reference_only_snapshot(
    snapshot: BitcoinEtfIssuerSnapshot,
    *,
    source_url: str,
) -> BitcoinEtfIssuerSnapshot:
    from dataclasses import replace as _replace

    return _replace(
        snapshot,
        source_url=source_url,
        source_type="aggregator",
        quality_status="degraded",
        collected_at=datetime.now(timezone.utc),
    )


def _candidate_structured_links(
    ticker: str, page_text: str, *, page_url: str
) -> list[tuple[str, str]]:
    links: list[tuple[str, str]] = []
    if ticker == "GBTC":
        links.append(("official_csv", GBTC_XLSX_URL))
    elif ticker == "BTC":
        links.append(("official_csv", BTC_MINI_XLSX_URL))
    discovered = _structured_download_links(page_text, base_url=page_url)
    for link in discovered:
        if link not in links:
            links.append(link)
    return links


def _ordered_structured_candidates(
    ticker: str,
    page_text: str,
    *,
    page_url: str,
) -> list[tuple[str, str]]:
    available_by_type: dict[str, list[str]] = {}
    for source_type, link in _candidate_structured_links(ticker, page_text, page_url=page_url):
        available_by_type.setdefault(source_type, [])
        if link not in available_by_type[source_type]:
            available_by_type[source_type].append(link)

    ordered: list[tuple[str, str]] = []
    for source_type in STRUCTURED_SOURCE_PRIORITY.get(ticker, ()):
        if source_type == HTML_ONLY_SOURCE_TYPE:
            continue
        for link in available_by_type.get(source_type, []):
            ordered.append((source_type, link))
    return ordered


def _collect_primary_snapshot(
    *,
    ticker: str,
    issuer: str,
    page_url: str,
    html_parser,
) -> CollectedOfficialSnapshot:
    def _try_structured_candidates(
        candidates: list[tuple[str, str]],
    ) -> CollectedOfficialSnapshot | None:
        for source_type, source_url in candidates:
            try:
                return _collect_structured_snapshot(
                    ticker=ticker,
                    issuer=issuer,
                    page_url=page_url,
                    source_type=source_type,
                    source_url=source_url,
                )
            except Exception as exc:
                log_structured(
                    logger,
                    event="etf.structured_source_failed",
                    message="구조화 공식 소스를 읽지 못해 다음 후보로 넘어갑니다.",
                    level=logging.WARNING,
                    provider=OFFICIAL_BTC_ETF_PROVIDER,
                    ticker=ticker,
                    source_type=source_type,
                    source_url=source_url,
                    reason=str(exc),
                )
        return None

    direct_candidates = _ordered_structured_candidates(ticker, "", page_url=page_url)
    direct_result = _try_structured_candidates(direct_candidates)
    if direct_result is not None:
        return direct_result

    page_payload = get_bytes_with_retry(
        page_url,
        provider=OFFICIAL_BTC_ETF_PROVIDER,
        headers={"Accept": "text/html,application/xhtml+xml"},
    )
    page_text = page_payload.decode("utf-8", errors="ignore")

    structured_candidates = [
        candidate
        for candidate in _ordered_structured_candidates(ticker, page_text, page_url=page_url)
        if candidate not in direct_candidates
    ]
    structured_result = _try_structured_candidates(structured_candidates)
    if structured_result is not None:
        return structured_result

    snapshot = html_parser(page_text)
    return CollectedOfficialSnapshot(
        snapshot=snapshot,
        raw_payload=page_payload,
        source_format="html",
        parse_method=f"{ticker.lower()}_html",
        http_status=200,
    )


def _fetch_direct_reference_snapshots(
    api_key: str = "",
    *,
    observer: PipelineObserver | None = None,
) -> list[BitcoinEtfIssuerSnapshot]:
    run_id = _current_run_id(observer)
    bundle = build_storage_bundle_from_env()
    snapshots: list[BitcoinEtfIssuerSnapshot] = []
    failures: list[dict[str, str]] = []
    reference_only_count = 0
    targets = (
        ("IBIT", "iShares", IBIT_URL, parse_ibit_snapshot),
        ("BITB", "Bitwise", BITB_URL, parse_bitb_snapshot),
        ("GBTC", "Grayscale", GBTC_URL, parse_gbtc_snapshot),
        ("BTC", "Grayscale Bitcoin Mini Trust", BTC_MINI_URL, parse_btc_mini_snapshot),
    )

    missing_tickers: set[str] = set()
    for ticker, issuer, url, parser in targets:
        try:
            collected = _collect_primary_snapshot(
                ticker=ticker,
                issuer=issuer,
                page_url=url,
                html_parser=parser,
            )
            snapshots.append(collected.snapshot)
            _persist_collected_snapshot(bundle, run_id=run_id, collected=collected)
        except Exception as exc:
            log_structured(
                logger,
                event="etf.primary_source_failed",
                message="공식 발행사 페이지에서 primary 스냅샷을 만들지 못했어요.",
                level=logging.WARNING,
                provider=OFFICIAL_BTC_ETF_PROVIDER,
                ticker=ticker,
                reason=str(exc),
                error_type=type(exc).__name__,
            )
            failures.append({"ticker": ticker, "detail": str(exc)})
            missing_tickers.add(ticker)

    if api_key and missing_tickers:
        try:
            reference_snapshots = _request_reference_snapshots(api_key, observer=observer)
        except Exception as exc:
            log_structured(
                logger,
                event="etf.reference_fetch_failed",
                message="reference-only 보조 소스 수집에 실패했어요.",
                level=logging.WARNING,
                provider=PERPLEXITY_PROVIDER,
                reason=str(exc),
            )
            reference_snapshots = []

        for snapshot in reference_snapshots:
            if snapshot.ticker not in missing_tickers:
                continue
            reference_snapshot = _reference_only_snapshot(snapshot, source_url=snapshot.source_url)
            collected = CollectedOfficialSnapshot(
                snapshot=reference_snapshot,
                raw_payload=json.dumps(asdict(reference_snapshot), cls=_DateTimeEncoder).encode(
                    "utf-8"
                ),
                source_format="json",
                parse_method="reference_snapshot",
                http_status=200,
                reference_only=True,
            )
            _persist_collected_snapshot(bundle, run_id=run_id, collected=collected)
            reference_only_count += 1
            log_structured(
                logger,
                event="etf.reference_only_snapshot",
                message="reference-only ETF 스냅샷을 별도로 기록했어요.",
                level=logging.INFO,
                ticker=reference_snapshot.ticker,
                source_url=reference_snapshot.source_url,
            )
            if observer is not None:
                observer.log_event(
                    "etf.reference_only_snapshot",
                    ticker=reference_snapshot.ticker,
                    source_url=reference_snapshot.source_url,
                )

    snapshots.sort(key=lambda item: item.ticker)
    quality_counts: dict[str, int] = {}
    for snapshot in snapshots:
        quality_counts[snapshot.quality_status] = quality_counts.get(snapshot.quality_status, 0) + 1
    log_structured(
        logger,
        event="etf.collection_quality",
        message="ETF 수집 품질 상태를 집계했어요.",
        run_id=run_id,
        quality_counts=quality_counts,
    )
    log_structured(
        logger,
        event="etf.run_summary",
        message="ETF 수집 실행 요약을 남겼어요.",
        run_id=run_id,
        primary_count=len(snapshots),
        reference_only_count=reference_only_count,
        failure_count=len(failures),
    )
    if observer is not None:
        observer.log_event("etf.collection_quality", run_id=run_id, quality_counts=quality_counts)
        observer.log_event(
            "etf.run_summary",
            run_id=run_id,
            snapshot_count=len(snapshots),
            reference_only_count=reference_only_count,
            failures=failures[:3],
        )
    return snapshots


def _build_client(api_key: str) -> Perplexity:
    return Perplexity(api_key=api_key, timeout=25, max_retries=1)


def _format_status_error(exc: APIStatusError) -> str:
    status_code = getattr(exc, "status_code", "unknown")
    response = getattr(exc, "response", None)
    detail = ""
    if response is not None:
        try:
            detail = str(response.text).strip()
        except Exception:
            detail = ""
    if detail:
        detail = " ".join(detail.split())[:240]
        return f"status={status_code}, detail={detail}"
    return f"status={status_code}"


def _retry_after_seconds_from_exception(exc: Exception) -> float | None:
    response = getattr(exc, "response", None)
    headers = getattr(response, "headers", None)
    if isinstance(headers, dict):
        return parse_retry_after_seconds(headers.get("Retry-After") or headers.get("retry-after"))
    return None


def _to_http_fetch_error(exc: Exception) -> HttpFetchError:
    if isinstance(exc, RateLimitError):
        message = f"Perplexity ETF 참조 요청 한도에 걸렸어요: {_format_status_error(exc)}"
        open_circuit(PERPLEXITY_PROVIDER, message)
        return HttpFetchError(
            message,
            provider=PERPLEXITY_PROVIDER,
            retryable=False,
            rate_limited=True,
            retry_after_seconds=_retry_after_seconds_from_exception(exc),
        )

    if isinstance(exc, APITimeoutError):
        return HttpFetchError(
            "Perplexity ETF 참조 응답 시간이 너무 오래 걸렸어요.",
            provider=PERPLEXITY_PROVIDER,
            retryable=True,
        )

    if isinstance(exc, APIConnectionError):
        return HttpFetchError(
            "Perplexity ETF 참조 연결을 열지 못했어요.",
            provider=PERPLEXITY_PROVIDER,
            retryable=True,
        )

    if isinstance(exc, APIStatusError):
        status_code = getattr(exc, "status_code", None)
        retry_after_seconds = _retry_after_seconds_from_exception(exc)
        if status_code == 429:
            message = f"Perplexity ETF 참조 요청 한도에 걸렸어요: {_format_status_error(exc)}"
            open_circuit(PERPLEXITY_PROVIDER, message)
            return HttpFetchError(
                message,
                provider=PERPLEXITY_PROVIDER,
                retryable=False,
                rate_limited=True,
                retry_after_seconds=retry_after_seconds,
            )
        return HttpFetchError(
            f"Perplexity ETF 참조 요청이 거절됐어요: {_format_status_error(exc)}",
            provider=PERPLEXITY_PROVIDER,
            retryable=status_code in policy_for(PERPLEXITY_PROVIDER).retryable_statuses,
            retry_after_seconds=retry_after_seconds,
        )

    return HttpFetchError(
        f"Perplexity ETF 참조 데이터를 가져오지 못했어요: {exc}",
        provider=PERPLEXITY_PROVIDER,
        retryable=False,
    )


def _extract_response_text(payload: dict[str, Any]) -> str:
    output_text = payload.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()

    choice_text = _extract_choice_message_text(payload.get("choices"))
    if choice_text:
        return choice_text

    raise HttpFetchError("Perplexity ETF 참조 응답에서 JSON 본문을 찾지 못했어요.")


def _extract_choice_message_text(choices: object) -> str:
    if not isinstance(choices, list) or not choices:
        return ""
    first = choices[0]
    if not isinstance(first, dict):
        return ""
    message = first.get("message")
    if not isinstance(message, dict):
        return ""
    content = message.get("content")
    if isinstance(content, str) and content.strip():
        return content.strip()
    if not isinstance(content, list):
        return ""

    parts = [
        text
        for item in content
        if isinstance(item, dict)
        for text in [str(item.get("text", "")).strip()]
        if text
    ]
    return "\n".join(parts)


def _decode_reference_snapshot_json(text: str) -> dict[str, Any]:
    stripped = text.strip().lstrip("\ufeff")
    if not stripped:
        raise HttpFetchError("Perplexity ETF 참조 응답에서 JSON 본문을 찾지 못했어요.")

    seen: set[str] = set()
    pending_candidates = _json_candidate_strings(stripped)
    while pending_candidates:
        candidate = pending_candidates.pop(0)
        normalized_candidate = candidate.strip()
        if not normalized_candidate or normalized_candidate in seen:
            continue
        seen.add(normalized_candidate)
        try:
            decoded = json.loads(normalized_candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(decoded, str):
            _append_nested_json_candidates(decoded, pending_candidates, seen)
            continue
        found = _find_snapshot_payload(decoded)
        if found is not None:
            return found

    raise HttpFetchError(
        f"Perplexity ETF 참조 JSON을 읽지 못했어요. preview={_response_preview(stripped)}"
    )


def _json_candidate_strings(text: str) -> list[str]:
    candidates: list[str] = []
    code_block_match = JSON_CODE_BLOCK_RE.search(text)
    if code_block_match:
        candidates.append(code_block_match.group("payload").strip())

    candidates.append(text)

    snapshot_member_candidate = _snapshot_member_candidate(text)
    if snapshot_member_candidate:
        candidates.append(snapshot_member_candidate)

    first_brace = text.find("{")
    last_brace = text.rfind("}")
    if first_brace != -1 and last_brace != -1 and first_brace < last_brace:
        candidates.append(text[first_brace : last_brace + 1].strip())

    return candidates


def _snapshot_member_candidate(text: str) -> str | None:
    snapshot_member_index = text.find('"snapshots"')
    if snapshot_member_index == -1:
        return None
    member_payload = text[snapshot_member_index:].strip().strip(",")
    return f"{{{member_payload}}}"


def _append_nested_json_candidates(
    decoded: str,
    pending_candidates: list[str],
    seen: set[str],
) -> None:
    nested = decoded.strip()
    if not nested or nested in seen:
        return
    pending_candidates.append(nested)
    snapshot_member_candidate = _snapshot_member_candidate(nested)
    if snapshot_member_candidate:
        pending_candidates.append(snapshot_member_candidate)


def _is_snapshot_payload(decoded: object) -> bool:
    return isinstance(decoded, dict) and (
        "snapshots" in decoded
        or {"ticker", "source_url", "shares_outstanding", "total_btc"} <= decoded.keys()
    )


def _find_snapshot_payload(decoded: object) -> dict[str, Any] | None:
    if _is_snapshot_payload(decoded):
        assert isinstance(decoded, dict)
        return decoded
    if isinstance(decoded, dict):
        for value in decoded.values():
            found = _find_snapshot_payload(value)
            if found is not None:
                return found
    if isinstance(decoded, list):
        for value in decoded:
            found = _find_snapshot_payload(value)
            if found is not None:
                return found
    return None


def _response_preview(text: str) -> str:
    preview = " ".join(str(text or "").split())
    if len(preview) <= RESPONSE_PREVIEW_LEN:
        return preview
    return f"{preview[:RESPONSE_PREVIEW_LEN]}..."


def _parse_numeric(value: object, *, integer: bool = False) -> int | float:
    if isinstance(value, (int, float)):
        return int(value) if integer else float(value)

    normalized = str(value or "").strip().replace(",", "").replace("$", "")
    if not normalized:
        raise ValueError("값이 비어 있어요.")
    numeric = float(normalized)
    return int(round(numeric)) if integer else numeric


def _is_allowed_official_url(url: str) -> bool:
    domain = normalize_domain(url)
    return any(domain_matches(domain, candidate) for candidate in BTC_ETF_REFERENCE_DOMAINS)


def _snapshot_from_item(item: dict[str, Any]) -> BitcoinEtfIssuerSnapshot:
    ticker = str(item.get("ticker", "")).strip().upper()
    source_url = str(item.get("source_url", "")).strip()
    as_of_raw = str(item.get("as_of", "")).strip()
    issuer = str(item.get("issuer", "")).strip() or ticker
    # Req 1.4: 공식 도메인 외부 URL이면 ValueError로 수집 중단
    if not ticker or not source_url or not as_of_raw or not _is_allowed_official_url(source_url):
        raise ValueError("필수 ETF 참조 필드가 없거나 공식 도메인이 아니에요.")

    try:
        as_of_date = _parse_as_of_date(as_of_raw)
    except ValueError as exc:
        raise ValueError(f"as_of 날짜 파싱 실패: {exc}") from exc

    shares_outstanding = int(_parse_numeric(item.get("shares_outstanding"), integer=True))
    daily_volume = int(_parse_numeric(item.get("daily_volume"), integer=True))
    aum_usd = float(_parse_numeric(item.get("aum_usd")))
    total_btc = float(_parse_numeric(item.get("total_btc")))
    bitcoin_per_share_raw = item.get("bitcoin_per_share")
    if bitcoin_per_share_raw in (None, ""):
        bitcoin_per_share = total_btc / shares_outstanding
    else:
        bitcoin_per_share = float(_parse_numeric(bitcoin_per_share_raw))

    if shares_outstanding <= 0 or daily_volume < 0 or aum_usd <= 0 or total_btc <= 0:
        raise ValueError("ETF 참조 수치가 유효 범위를 벗어나요.")

    snapshot = BitcoinEtfIssuerSnapshot(
        ticker=ticker,
        issuer=issuer,
        source_url=source_url,
        as_of_date=as_of_date,
        shares_outstanding=shares_outstanding,
        daily_volume=daily_volume,
        aum_usd=round(aum_usd, 2),
        total_btc=round(total_btc, 8),
        bitcoin_per_share=round(bitcoin_per_share, 10),
        source_type="aggregator",
        quality_status="degraded",
    )
    return _validate_snapshot_anomalies(snapshot)


def _parse_reference_snapshot_text(text: str) -> list[BitcoinEtfIssuerSnapshot]:
    decoded = _decode_reference_snapshot_json(text)
    if not isinstance(decoded, dict):
        raise HttpFetchError("Perplexity ETF 참조 JSON 구조가 예상과 달라요.")

    if (
        "snapshots" not in decoded
        and {
            "ticker",
            "source_url",
            "shares_outstanding",
            "total_btc",
        }
        <= decoded.keys()
    ):
        decoded = {"snapshots": [decoded]}

    raw_snapshots = decoded.get("snapshots", [])
    if not isinstance(raw_snapshots, list):
        raise HttpFetchError("Perplexity ETF 참조 JSON에 snapshots 배열이 없어요.")

    snapshots: list[BitcoinEtfIssuerSnapshot] = []
    for raw_item in raw_snapshots:
        if not isinstance(raw_item, dict):
            continue
        try:
            snapshots.append(_snapshot_from_item(raw_item))
        except ValueError as exc:
            log_structured(
                logger,
                event="selection.complete",
                message="Perplexity ETF 참조 항목을 건너뛸게요.",
                level=logging.WARNING,
                provider=PERPLEXITY_PROVIDER,
                reason=str(exc),
                kept_count=0,
            )

    snapshots.sort(key=lambda item: item.ticker)
    return snapshots


def _parse_reference_snapshot_response(payload: dict[str, Any]) -> list[BitcoinEtfIssuerSnapshot]:
    text = _extract_response_text(payload)
    return _parse_reference_snapshot_text(text)


def _request_reference_snapshots(
    api_key: str,
    *,
    observer: PipelineObserver | None = None,
) -> list[BitcoinEtfIssuerSnapshot]:
    if not api_key:
        log_structured(
            logger,
            event="phase.skip",
            message="Perplexity API 키가 없어 BTC ETF 참조 스냅샷은 건너뛸게요.",
            level=logging.DEBUG,
            provider=PERPLEXITY_PROVIDER,
            reason="missing_api_key",
        )
        return []

    unavailable_reason = disabled_reason(PERPLEXITY_PROVIDER)
    if unavailable_reason:
        record_skip(PERPLEXITY_PROVIDER)
        raise HttpFetchError(
            f"Perplexity는 이번 실행에서 더 이상 쓰지 않을게요: {unavailable_reason}"
        )

    client = _build_client(api_key)
    if observer is not None:
        observer.record_provider_usage(PERPLEXITY_PROVIDER, requests=1)

    def request_once() -> list[BitcoinEtfIssuerSnapshot]:
        try:
            response = client.chat.completions.create(
                model=BTC_ETF_REFERENCE_MODEL,
                messages=[
                    {
                        "role": "system",
                        "content": "You normalize financial reference data into strict JSON.",
                    },
                    {
                        "role": "user",
                        "content": _render_reference_prompt(date.today()),
                    },
                ],
                search_domain_filter=list(BTC_ETF_REFERENCE_DOMAINS),
                search_recency_filter="month",
                search_mode="web",
                web_search_options={
                    "user_location": {"country": "US"},
                },
                temperature=0.0,
                max_tokens=900,
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "btc_etf_snapshots",
                        "schema": BTC_ETF_REFERENCE_RESPONSE_SCHEMA,
                    },
                },
            )
        except (RateLimitError, APITimeoutError, APIConnectionError, APIStatusError) as exc:
            raise _to_http_fetch_error(exc) from exc

        try:
            payload = response.model_dump()
        except AttributeError:
            if isinstance(response, dict):
                payload = response
            else:
                raise HttpFetchError("Perplexity ETF 참조 응답 구조가 예상과 달라요.")

        if not isinstance(payload, dict):
            raise HttpFetchError("Perplexity ETF 참조 응답 구조가 예상과 달라요.")

        text = _extract_response_text(payload)
        snapshots = _parse_reference_snapshot_text(text)
        if observer is not None and not snapshots:
            observer.log_event(
                "btc_etf_reference_parse_empty",
                response_preview=_response_preview(text),
                source_domain_count=len(BTC_ETF_REFERENCE_DOMAINS),
            )
        return snapshots

    snapshots = execute_with_provider_retry(
        provider=PERPLEXITY_PROVIDER,
        operation=request_once,
        should_retry=lambda exc: isinstance(exc, HttpFetchError) and exc.retryable,
        on_retry=lambda exc, attempt, max_attempts, delay: log_structured(
            logger,
            event="provider.retry",
            message="Perplexity ETF 참조 데이터를 다시 시도하는 중이에요.",
            level=logging.WARNING,
            provider=PERPLEXITY_PROVIDER,
            attempt=attempt,
            max_attempts=max_attempts,
            reason=str(exc),
            retryable=True,
            delay_seconds=delay,
        ),
        retry_after_seconds_for_error=lambda exc: (
            exc.retry_after_seconds if isinstance(exc, HttpFetchError) else None
        ),
    )
    if observer is not None:
        observer.record_provider_usage(
            PERPLEXITY_PROVIDER,
            response_sources=len(snapshots),
        )
    return snapshots


def fetch_official_btc_etf_snapshots(
    *,
    api_key: str = "",
    observer: PipelineObserver | None = None,
) -> list[BitcoinEtfIssuerSnapshot]:
    return _fetch_direct_reference_snapshots(api_key=api_key, observer=observer)


class _DateTimeEncoder(json.JSONEncoder):
    """date/datetime을 ISO 문자열로 직렬화하는 JSON 인코더."""

    def default(self, obj: object) -> object:
        if isinstance(obj, datetime):
            return obj.isoformat()
        if isinstance(obj, date):
            return obj.isoformat()
        return super().default(obj)


def _deserialize_snapshot_item(item: dict[str, Any]) -> dict[str, Any]:
    """캐시에서 읽은 dict를 BitcoinEtfIssuerSnapshot(**item) 가능한 형태로 변환."""
    item = dict(item)
    # as_of (구버전) → as_of_date 마이그레이션
    if "as_of" in item and "as_of_date" not in item:
        try:
            item["as_of_date"] = _parse_as_of_date(item.pop("as_of"))
        except ValueError:
            item.pop("as_of", None)
            return {}
    elif "as_of_date" in item and isinstance(item["as_of_date"], str):
        try:
            item["as_of_date"] = date.fromisoformat(item["as_of_date"])
        except ValueError:
            return {}
    if "collected_at" in item and isinstance(item["collected_at"], str):
        try:
            item["collected_at"] = datetime.fromisoformat(item["collected_at"])
        except ValueError:
            item["collected_at"] = None
    return item


def load_official_btc_etf_cache(cache_file: Path) -> dict[str, BitcoinEtfIssuerSnapshot]:
    if not cache_file.exists():
        return {}

    try:
        payload = json.loads(cache_file.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        return {}

    if not isinstance(payload, dict):
        return {}

    snapshots: dict[str, BitcoinEtfIssuerSnapshot] = {}
    for ticker, item in payload.items():
        if not isinstance(item, dict):
            continue
        try:
            deserialized = _deserialize_snapshot_item(item)
            if not deserialized:
                continue
            snapshots[str(ticker)] = BitcoinEtfIssuerSnapshot(**deserialized)
        except TypeError:
            continue
    return snapshots


def save_official_btc_etf_cache(
    cache_file: Path, snapshots: list[BitcoinEtfIssuerSnapshot]
) -> None:
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    payload = {snapshot.ticker: asdict(snapshot) for snapshot in snapshots}
    cache_file.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, cls=_DateTimeEncoder),
        encoding="utf-8",
    )


def save_official_btc_etf_cache_state(
    cache_dir: Path,
    *,
    snapshot_count: int,
    reason: str,
) -> None:
    cache_dir.mkdir(parents=True, exist_ok=True)
    state_file = cache_dir / OFFICIAL_BTC_ETF_STATE_FILE
    payload = {
        "fetched_at_utc": datetime.now(timezone.utc).isoformat(),
        "snapshot_count": snapshot_count,
        "reason": reason,
    }
    state_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


__all__ = [
    "BitcoinEtfIssuerSnapshot",
    "BTC_MINI_URL",
    "GBTC_URL",
    "IBIT_URL",
    "BITB_URL",
    "fetch_official_btc_etf_snapshots",
    "load_official_btc_etf_cache",
    "parse_bitb_snapshot",
    "parse_btc_mini_snapshot",
    "parse_gbtc_snapshot",
    "parse_ibit_snapshot",
    "save_official_btc_etf_cache",
    "save_official_btc_etf_cache_state",
]
