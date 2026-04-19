#!/usr/bin/env python3
"""GBTC + IBIT ETF 보유량 이력을 Supabase btc_etf_gold 테이블에 백필합니다.

GBTC: Grayscale S3 XLSX (약 1,254행, 2021-04-19~)
IBIT: iShares 펀드 XLS  (약  574행, 2024-01-05~)

필요 환경변수:
    SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY

선택 환경변수:
    BTC_ETF_GOLD_TABLE (기본값: btc_etf_gold)

실행:
    python scripts/backfill_btc_etf_history.py
    python scripts/backfill_btc_etf_history.py --dry-run
"""

from __future__ import annotations

import argparse
import io
import logging
import os
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from morning_brief.data.etf_storage import DEFAULT_ETF_GOLD_TABLE, build_gold_record
from morning_brief.data.sources.http_client import get_bytes_with_retry
from morning_brief.models import BitcoinEtfIssuerSnapshot

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

GBTC_XLSX_URL = (
    "https://reporting-prod-20231113144948145500000003.s3.us-east-1.amazonaws.com"
    "/product-performance/672e88c7-dac6-4fcd-9069-18eef01a2c73.xlsx"
)
IBIT_PAGE_URL = "https://www.ishares.com/us/products/333011/ishares-bitcoin-trust-etf"
# iShares fund XLS (historical NAV + shares) — product-specific static ID
IBIT_XLS_SUFFIX = "/1521942788811.ajax?fileType=xls"
# iShares holdings CSV (current snapshot — used for btc_per_share)
IBIT_HOLDINGS_SUFFIX = "/1467271812596.ajax?fileType=csv"

PROVIDER = "btc_etf_backfill"
RUN_ID = f"backfill-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"


# ---------------------------------------------------------------------------
# XLSX 파싱 유틸
# ---------------------------------------------------------------------------


def _read_xlsx_sheets(payload: bytes) -> dict[str, list[list[str]]]:
    """XLSX / XLS / XML Spreadsheet 2003 bytes → {sheet_name: [[cell, ...], ...]}.

    시도 순서:
      1. openpyxl  — 진짜 XLSX
      2. zipfile + ElementTree  — XLSX (ZIP) 직접 파싱
      3. ElementTree  — XML Spreadsheet 2003 (iShares XLS 포맷: ss: namespace)
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        result: dict[str, list[list[str]]] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            result[name] = [
                [("" if v is None else str(v).strip()) for v in row]
                for row in ws.iter_rows(values_only=True)
            ]
        return result
    except Exception:
        pass

    # fallback 2: stdlib zipfile + ElementTree (진짜 XLSX = ZIP 컨테이너)
    try:
        import zipfile
        from xml.etree import ElementTree as ET

        archive = zipfile.ZipFile(io.BytesIO(payload))
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            shared = ["".join(n.itertext()).strip() for n in root.findall(f"{ns}si")]

        def _cell_val(cell: ET.Element) -> str:
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            v = cell.find(f"{ns}v")
            if v is None or v.text is None:
                return ""
            if cell.get("t") == "s":
                try:
                    return shared[int(v.text)]
                except (ValueError, IndexError):
                    return ""
            return v.text.strip()

        wb_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        r_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
        rel_map = {r.attrib.get("Id", ""): r.attrib.get("Target", "") for r in rels_root}
        sheets_out: dict[str, list[list[str]]] = {}
        for s in wb_root.findall(f"{ns}sheets/{ns}sheet"):
            sname = s.attrib.get("name", "")
            rid = s.attrib.get(f"{r_ns}id", "")
            target = rel_map.get(rid, "")
            if not sname or not target:
                continue
            path = target if target.startswith("xl/") else f"xl/{target}"
            if path not in archive.namelist():
                continue
            sroot = ET.fromstring(archive.read(path))
            col_re = re.compile(r"^([A-Z]+)")
            rows_out: list[list[str]] = []
            for row_el in sroot.findall(f".//{ns}row"):
                cells: list[tuple[int, str]] = []
                for c in row_el.findall(f"{ns}c"):
                    m = col_re.match(c.attrib.get("r", "A"))
                    if not m:
                        continue
                    col_idx = (
                        sum((ord(ch) - 64) * (26**i) for i, ch in enumerate(reversed(m.group(1))))
                        - 1
                    )
                    cells.append((col_idx, _cell_val(c)))
                if cells:
                    max_col = max(idx for idx, _ in cells) + 1
                    row_data = [""] * max_col
                    for idx, val in cells:
                        row_data[idx] = val
                    rows_out.append(row_data)
            sheets_out[sname] = rows_out
        if sheets_out:
            return sheets_out
    except Exception:
        pass

    # fallback 3: XML Spreadsheet 2003 (iShares XLS — ss: namespace)
    try:
        from xml.etree import ElementTree as _ET2

        _NS = "urn:schemas-microsoft-com:office:spreadsheet"
        content = payload.lstrip(b"\xef\xbb\xbf")  # strip UTF-8 BOM(s)
        # URL 속 & 가 &amp; 미처리된 well-formed 위반을 교정합니다.
        content = re.sub(rb"&(?!amp;|lt;|gt;|quot;|apos;|#)", rb"&amp;", content)
        root = _ET2.fromstring(content)
        sheets_ss: dict[str, list[list[str]]] = {}
        for ws in root.findall(f"{{{_NS}}}Worksheet"):
            sname = ws.get(f"{{{_NS}}}Name", "Sheet")
            table = ws.find(f"{{{_NS}}}Table")
            if table is None:
                continue
            rows_ss: list[list[str]] = []
            for row_el in table.findall(f"{{{_NS}}}Row"):
                cell_map: dict[int, str] = {}
                col_idx = 0
                for cell_el in row_el.findall(f"{{{_NS}}}Cell"):
                    explicit = cell_el.get(f"{{{_NS}}}Index")
                    if explicit:
                        col_idx = int(explicit) - 1
                    data_el = cell_el.find(f"{{{_NS}}}Data")
                    val = (data_el.text or "").strip() if data_el is not None else ""
                    cell_map[col_idx] = val
                    col_idx += 1
                if cell_map:
                    max_col = max(cell_map) + 1
                    row_data = [""] * max_col
                    for idx, val in cell_map.items():
                        row_data[idx] = val
                    rows_ss.append(row_data)
            sheets_ss[sname] = rows_ss
        return sheets_ss
    except Exception:
        pass

    return sheets_out


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def _parse_float(s: str) -> float | None:
    try:
        return float(s.replace(",", "").replace("$", "").strip())
    except (ValueError, AttributeError):
        return None


def _parse_date_str(s: str) -> date | None:
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%b %d, %Y", "%d-%b-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Excel serial number (float string)
    try:
        serial = float(s)
        if 30000 < serial < 60000:
            from datetime import timedelta

            return date(1899, 12, 30) + timedelta(days=int(serial))
    except ValueError:
        pass
    return None


# ---------------------------------------------------------------------------
# GBTC 파싱
# ---------------------------------------------------------------------------


def _gbtc_snapshots(payload: bytes) -> list[BitcoinEtfIssuerSnapshot]:
    sheets = _read_xlsx_sheets(payload)
    if not sheets:
        logger.error("GBTC XLSX 파싱 실패 — sheets 없음")
        return []

    # Holdings 시트에서 btc_per_share 추출
    btc_per_share = 0.00077796  # 하드코딩 fallback
    for sname, rows in sheets.items():
        if "hold" in sname.lower():
            for row in rows:
                for i, cell in enumerate(row):
                    if "asset" in cell.lower() and "share" in cell.lower():
                        # 오른쪽 인접 셀에 값 있음
                        for j in range(i + 1, min(i + 4, len(row))):
                            v = _parse_float(row[j])
                            if v is not None and 0.0001 < v < 0.01:
                                btc_per_share = v
                                break
    logger.info("GBTC btc_per_share=%.8f", btc_per_share)

    # Daily Performance 시트 파싱
    for sname, rows in sheets.items():
        if "daily" not in sname.lower() and "performance" not in sname.lower():
            continue

        # 헤더 행 탐색
        header_idx = -1
        headers: list[str] = []
        for i, row in enumerate(rows):
            norms = [_norm(c) for c in row]
            if "date" in norms and "sharesoutstanding" in norms:
                header_idx = i
                headers = norms
                break
        if header_idx == -1:
            logger.warning("GBTC Daily Performance 헤더 미발견 — 시트=%s", sname)
            continue

        snapshots: list[BitcoinEtfIssuerSnapshot] = []
        for row in rows[header_idx + 1 :]:
            if len(row) < len(headers):
                row = row + [""] * (len(headers) - len(row))
            record = {headers[j]: row[j] for j in range(len(headers)) if headers[j]}

            date_raw = record.get("date", "")
            shares_raw = record.get("sharesoutstanding", "")
            nav_raw = record.get("navpershare", "")
            aum_raw = record.get("aum", "")

            as_of = _parse_date_str(date_raw)
            shares = _parse_float(shares_raw)
            nav = _parse_float(nav_raw)
            if as_of is None or shares is None or nav is None:
                continue

            shares_int = max(int(shares), 0)
            total_btc = round(shares_int * btc_per_share, 8)
            aum = _parse_float(aum_raw) if aum_raw else None
            if aum is None or aum <= 0:
                aum = round(nav * shares_int, 2)

            snapshots.append(
                BitcoinEtfIssuerSnapshot(
                    ticker="GBTC",
                    issuer="Grayscale",
                    source_url=GBTC_XLSX_URL,
                    as_of_date=as_of,
                    shares_outstanding=shares_int,
                    daily_volume=0,
                    aum_usd=aum,
                    total_btc=total_btc,
                    bitcoin_per_share=btc_per_share,
                    source_type="official_csv",
                    quality_status="ok",
                    collected_at=datetime.now(timezone.utc),
                )
            )
        logger.info("GBTC 파싱 완료 — %d행", len(snapshots))
        return snapshots

    logger.error("GBTC Daily Performance 시트 미발견")
    return []


# ---------------------------------------------------------------------------
# IBIT 파싱
# ---------------------------------------------------------------------------


def _discover_ibit_xls_url() -> str | None:
    """iShares 제품 페이지에서 XLS(historical NAV) 다운로드 URL을 탐색합니다."""
    try:
        page = get_bytes_with_retry(IBIT_PAGE_URL, provider=PROVIDER, timeout=30)
        text = page.decode("utf-8", errors="replace")
        # 패턴: /products/333011/.../NNNNNNNNNNNN.ajax?fileType=xls
        m = re.search(r"(/us/products/333011/[^\"']+?/(\d{10,}).ajax\?fileType=xls)", text)
        if m:
            return "https://www.ishares.com" + m.group(1)
    except Exception as exc:
        logger.warning("IBIT 페이지 탐색 실패: %s", exc)
    # 알려진 정적 URL 사용
    return (
        "https://www.ishares.com"
        + IBIT_PAGE_URL.replace("https://www.ishares.com", "").rstrip("/")
        + IBIT_XLS_SUFFIX
    )


def _ibit_btc_per_share() -> float:
    """iShares Holdings CSV에서 현재 total_btc를 추출합니다.

    CSV 앞쪽에 메타데이터 행(펀드명, Inception Date 등)이 있으므로
    실제 컬럼 헤더 행(Ticker,Name,...)을 먼저 찾아서 파싱합니다.
    """
    try:
        import csv

        url = IBIT_PAGE_URL.rstrip("/") + IBIT_HOLDINGS_SUFFIX
        payload = get_bytes_with_retry(url, provider=PROVIDER, timeout=30)
        text = payload.decode("utf-8-sig", errors="replace")
        lines = text.splitlines()
        # 실제 헤더 행: "Ticker" 로 시작하는 첫 번째 행
        header_start = next(
            (i for i, line in enumerate(lines) if line.strip().startswith("Ticker")), None
        )
        if header_start is None:
            logger.warning("IBIT Holdings CSV — Ticker 헤더 행 미발견")
            return 0.0
        reader = csv.DictReader(io.StringIO("\n".join(lines[header_start:])))
        total_btc: float | None = None
        for row in reader:
            name = row.get("Name", "")
            if "bitcoin" in name.lower() or "btc" in name.lower():
                qty = row.get("Quantity", "")
                v = _parse_float(qty)
                if v and v > 0:
                    total_btc = v
                    break
        if total_btc:
            logger.info("IBIT holdings CSV total_btc=%.2f", total_btc)
        return total_btc or 0.0
    except Exception as exc:
        logger.warning("IBIT Holdings CSV 수집 실패: %s", exc)
        return 0.0


def _ibit_snapshots(payload: bytes, ibit_total_btc_now: float) -> list[BitcoinEtfIssuerSnapshot]:
    """iShares 펀드 XLS에서 전체 이력 스냅샷을 파싱합니다."""
    sheets = _read_xlsx_sheets(payload)
    if not sheets:
        logger.error("IBIT XLS 파싱 실패 — sheets 없음")
        return []

    snapshots: list[BitcoinEtfIssuerSnapshot] = []
    for sname, rows in sheets.items():
        # 헤더 탐색: "As Of" + "Shares Outstanding" 조합
        header_idx = -1
        headers: list[str] = []
        for i, row in enumerate(rows):
            norms = [_norm(c) for c in row]
            has_date = any(k in norms for k in ("asof", "date", "navdate"))
            has_shares = "sharesoutstanding" in norms
            if has_date and has_shares:
                header_idx = i
                headers = norms
                break
        if header_idx == -1:
            continue

        date_key = next((k for k in ("asof", "navdate", "date") if k in headers), None)
        shares_key = "sharesoutstanding" if "sharesoutstanding" in headers else None
        nav_key = "navpershare" if "navpershare" in headers else None
        if not date_key or not shares_key:
            continue

        # Historical 시트는 newest-first → 첫 번째 유효 행이 최신 shares
        latest_shares: float | None = None
        for row in rows[header_idx + 1 :]:
            if len(row) < len(headers):
                row = row + [""] * (len(headers) - len(row))
            record = {headers[j]: row[j] for j in range(len(headers))}
            s = _parse_float(record.get(shares_key, ""))
            if s and s > 0:
                latest_shares = s
                break

        if not latest_shares or not ibit_total_btc_now:
            logger.warning(
                "IBIT btc_per_share 계산 불가 — latest_shares=%s, total_btc=%s",
                latest_shares,
                ibit_total_btc_now,
            )
            # 최근 실측값 fallback (2026년 4월 기준 IBIT btc/share ≈ 0.000568)
            btc_per_share = 0.000568
        else:
            btc_per_share = ibit_total_btc_now / latest_shares
        logger.info(
            "IBIT btc_per_share=%.8f (latest_shares=%.0f, total_btc=%.2f)",
            btc_per_share,
            latest_shares or 0,
            ibit_total_btc_now,
        )

        for row in rows[header_idx + 1 :]:
            if len(row) < len(headers):
                row = row + [""] * (len(headers) - len(row))
            record = {headers[j]: row[j] for j in range(len(headers))}

            as_of = _parse_date_str(record.get(date_key, ""))
            shares = _parse_float(record.get(shares_key, ""))
            nav = _parse_float(record.get(nav_key, "")) if nav_key else None

            if as_of is None or shares is None or shares <= 0:
                continue

            shares_int = int(shares)
            total_btc = round(shares_int * btc_per_share, 8)
            aum = round((nav or 0.0) * shares_int, 2) if nav else None

            snapshots.append(
                BitcoinEtfIssuerSnapshot(
                    ticker="IBIT",
                    issuer="iShares",
                    source_url=IBIT_PAGE_URL,
                    as_of_date=as_of,
                    shares_outstanding=shares_int,
                    daily_volume=0,
                    aum_usd=aum or 0.0,
                    total_btc=total_btc,
                    bitcoin_per_share=btc_per_share,
                    source_type="official_csv",
                    quality_status="ok",
                    collected_at=datetime.now(timezone.utc),
                )
            )
        logger.info("IBIT 파싱 완료 — %d행 (시트=%s)", len(snapshots), sname)
        return snapshots

    logger.error("IBIT에서 헤더 행(AsOf + SharesOutstanding) 미발견")
    return []


# ---------------------------------------------------------------------------
# Supabase upsert
# ---------------------------------------------------------------------------


def _upsert_gold(
    snapshots: list[BitcoinEtfIssuerSnapshot],
    *,
    table: str,
    dry_run: bool,
) -> int:
    if not snapshots:
        return 0

    records: list[dict[str, Any]] = []
    for s in snapshots:
        rec = build_gold_record(s, run_id=RUN_ID)
        rec["bronze_object_path"] = "backfill"
        records.append(rec)

    if dry_run:
        logger.info("[dry-run] upsert 대상 %d건 (table=%s)", len(records), table)
        for r in records[:3]:
            logger.info(
                "  sample: ticker=%s as_of_date=%s total_btc=%.2f",
                r["ticker"],
                r["as_of_date"],
                r.get("total_btc", 0),
            )
        return len(records)

    supabase_url = os.getenv("SUPABASE_URL", "").strip()
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not supabase_url or not service_role_key:
        logger.error("SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY 환경변수 미설정")
        return 0

    from supabase import create_client

    client = create_client(supabase_url, service_role_key)
    BATCH = 200
    upserted = 0
    for i in range(0, len(records), BATCH):
        batch = records[i : i + BATCH]
        client.table(table).upsert(batch, on_conflict="ticker,as_of_date").execute()
        upserted += len(batch)
        logger.info("  upsert %d/%d (table=%s)", upserted, len(records), table)
    return upserted


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description="BTC ETF 이력 백필")
    parser.add_argument("--dry-run", action="store_true", help="Supabase에 실제로 쓰지 않음")
    parser.add_argument("--ticker", choices=["GBTC", "IBIT", "all"], default="all")
    args = parser.parse_args()

    gold_table = (
        os.getenv("BTC_ETF_GOLD_TABLE", DEFAULT_ETF_GOLD_TABLE).strip() or DEFAULT_ETF_GOLD_TABLE
    )
    total = 0

    if args.ticker in ("GBTC", "all"):
        logger.info("=== GBTC XLSX 다운로드 중... ===")
        try:
            gbtc_payload = get_bytes_with_retry(GBTC_XLSX_URL, provider=PROVIDER, timeout=60)
            logger.info("다운로드 완료 (%d bytes)", len(gbtc_payload))
            gbtc_snapshots = _gbtc_snapshots(gbtc_payload)
            n = _upsert_gold(gbtc_snapshots, table=gold_table, dry_run=args.dry_run)
            logger.info("GBTC 완료 — %d행 upsert", n)
            total += n
        except Exception as exc:
            logger.error("GBTC 실패: %s", exc)

    if args.ticker in ("IBIT", "all"):
        logger.info("=== IBIT XLS 다운로드 중... ===")
        try:
            ibit_total_btc = _ibit_btc_per_share()
            xls_url = _discover_ibit_xls_url()
            logger.info("IBIT XLS URL: %s", xls_url)
            ibit_payload = get_bytes_with_retry(xls_url or "", provider=PROVIDER, timeout=60)
            logger.info("다운로드 완료 (%d bytes)", len(ibit_payload))
            ibit_snaps = _ibit_snapshots(ibit_payload, ibit_total_btc)
            n = _upsert_gold(ibit_snaps, table=gold_table, dry_run=args.dry_run)
            logger.info("IBIT 완료 — %d행 upsert", n)
            total += n
        except Exception as exc:
            logger.error("IBIT 실패: %s", exc)

    logger.info("=== 전체 완료 — 총 %d행 upsert ===", total)
    return 0


if __name__ == "__main__":
    sys.exit(main())
